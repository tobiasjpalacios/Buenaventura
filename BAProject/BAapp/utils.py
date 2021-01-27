from openpyxl import load_workbook, Workbook
from django.db import transaction

from .models import Articulo, Empresa

def sheet_reader(sheet):
    wb = load_workbook(filename=sheet)

    emp_sheet = wb["empresas"]
    emp_header = {}
    for i in range(len(emp_sheet[1])):
        key = emp_sheet[1][i].value.lower().replace(" ", "_")
        emp_header[key] = i
    with transaction.atomic():
        for c in emp_sheet.iter_rows(2, values_only=True):
            instance = Empresa()
            if (c[emp_header["id"]]):
                instance = Empresa.objects.get(id=c[emp_header["id"]])
            if (not c[emp_header["razon_social"]]):
                break
            for v in emp_header.keys():
                if (v=="retenciones"):
                    continue
                setattr(instance, v, c[emp_header[v]])
            instance.save()

    art_sheet = wb["articulos"]
    art_header = {}
    for i in range(len(art_sheet[1])):
        key = art_sheet[1][i].value.lower().replace(" ", "_")
        art_header[key] = i

    with transaction.atomic():
        for c in art_sheet.iter_rows(2, values_only=True):
            instance = Articulo()
            if (c[art_header["id"]]):
                instance = Articulo.objects.get(id=c[art_header["id"]])
            if (not c[art_header["marca"]]):
                break
            for v in art_header.keys():
                if (v=="empresa"):
                    continue
                setattr(instance, v, c[art_header[v]])
                instance.empresa = Empresa.objects.get(
                    razon_social=c[art_header["empresa"]])
            instance.save()


