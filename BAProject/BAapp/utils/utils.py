from io import BytesIO, FileIO

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname, get_column_letter
from django.db import IntegrityError, transaction
from django.db.models.fields.reverse_related import ManyToOneRel
from django.utils.safestring import mark_safe
from BAapp.models import MyUser

from BAapp.models import Articulo, Empresa, Retencion

def _read_header(ws):
    header = {}
    for i in range(len(ws[1])):
        key = ws[1][i].value.lower().replace(" ", "_")
        header[key] = i
    return header

def reed_empresas(emp_sheet):
    emp_header = _read_header(emp_sheet)
    emp_creadas = 0
    emp_actualizadas = 0
    for c in emp_sheet.iter_rows(2, values_only=True):
        instance = Empresa()
        if c[emp_header["id"]]:
            instance = Empresa.objects.get(id=c[emp_header["id"]])
            emp_actualizadas += 1
        if not c[emp_header["razon_social"]]:
            break
        for v in emp_header.keys():
            if v == "retenciones":
                if c[emp_header[v]]:    
                    rets = c[emp_header[v]].split("+")
                    for ret in rets:
                        obj = Retencion.objects.get(name=ret)
                        instance.retenciones.add(obj)
                continue
            setattr(instance, v, c[emp_header[v]])
        try:
            instance.save()
            emp_creadas += 1
        except IntegrityError:
            instance = Empresa.objects.get(razon_social=c[emp_header["razon_social"]])
            instance.nombre_comercial = c[emp_header["nombre_comercial"]]
            instance.cuit             = c[emp_header["cuit"]]
            instance.ingresos_brutos  = c[emp_header["ingresos_brutos"]]
            instance.fecha_exclusion  = c[emp_header["fecha_exclusion"]]
            instance.categoria_iva    = c[emp_header["categoria_iva"]]
            instance.domicilio_fiscal = c[emp_header["domicilio_fiscal"]]
            if c[emp_header["retenciones"]]:
                rets = c[emp_header["retenciones"]].split("+")
                for ret in rets:
                    obj = Retencion.objects.get(name=ret)
                    instance.retenciones.add(obj)
            instance.save()
            emp_actualizadas += 1
    
    return f"<h5>Empresas:</h5>Creadas: {emp_creadas} empresa(s)<br>Actualizadas: {emp_actualizadas} empresa(s)<br>"


def reed_articulos(art_sheet):
    art_header = _read_header(art_sheet)
    created_arts = 0
    for c in art_sheet.iter_rows(2, values_only=True):
        instance = Articulo()
        if c[art_header["id"]]:
            instance = Articulo.objects.get(id=c[art_header["id"]])
        if not c[art_header["marca"]]:
            break
        for v in art_header.keys():
            if v=="empresa":
                instance.empresa = Empresa.objects.get(
                    razon_social=c[art_header["empresa"]])
                continue
            setattr(instance, v, c[art_header[v]])
        # solucion temporal, mala y lenta (no se puede actualizar ningun objeto)
        emp_nombre_comercial = Empresa.objects.get(razon_social=c[art_header["empresa"]]).nombre_comercial
        art = Articulo.objects.filter(ingrediente=c[art_header["ingrediente"]], empresa__nombre_comercial=emp_nombre_comercial).exists()
        if not art:
            instance.save()
            created_arts += 1
    
    return f"<h5>Articulos:</h5>Creados: {created_arts} articulo(s)<br>"

def reed_usuarios(usr_sheet):
    created_users_count = 0
    updated_users_count = 0
    usr_reader = _read_header(usr_sheet)
    for i, row in enumerate(usr_sheet.iter_rows(2, values_only=True)):
        row_data = list()
        for cell in row:
            row_data.append(str(cell))
        if not row[usr_reader["email"]]:
            break
        conv = lambda el : "" if el == "None" or el == None else el
        user_data = [conv(d) for d in row_data]
        empresa = user_data[usr_reader["empresa"]]
        empresa = None if empresa == "None" or empresa == "" else Empresa.objects.get(razon_social=empresa)
        fech_nac = user_data[usr_reader["fecha_de_nacimiento"]]
        fech_nac = None if fech_nac == "None" or fech_nac == "" else fech_nac
        dni = user_data[usr_reader["dni"]]
        dni = None if dni == "None" or dni == "" else dni
        try:
            user = MyUser.objs.get(email=user_data[0])             
            user.email            = user_data[usr_reader["email"]]
            user.nombre           = user_data[usr_reader["nombre"]]
            user.apellido         = user_data[usr_reader["apellido"]]
            user.clase            = user_data[usr_reader["clase"]]
            user.empresa          = empresa
            user.fecha_nacimiento = fech_nac
            user.sexo             = user_data[usr_reader["sexo"]]
            user.dni              = dni
            user.telefono         = user_data[usr_reader["telefono"]]
            user.domicilio        = user_data[usr_reader["domicilio"]]
            user.save()
            updated_users_count += 1
        except MyUser.DoesNotExist:
            try:
                MyUser.objs.create_user(
                    email            = user_data[usr_reader["email"]],
                    nombre           = user_data[usr_reader["nombre"]],
                    apellido         = user_data[usr_reader["apellido"]],
                    password         = user_data[usr_reader["password"]],
                    clase            = user_data[usr_reader["clase"]],
                    empresa          = empresa,
                    fecha_nacimiento = fech_nac,
                    sexo             = user_data[usr_reader["sexo"]],
                    dni              = dni,
                    telefono         = user_data[usr_reader["telefono"]],
                    domicilio        = user_data[usr_reader["domicilio"]],
                )
                created_users_count += 1
            except Exception as e:
                return f"<h5>Usuarios:</h5><b>{e} (fila N°{i+2})</b><br>Creados: {created_users_count} usuario(s)<br>Actualizados: {updated_users_count} usuario(s)<br>"
    # end for
    return f"<h5>Usuarios:</h5>Creados: {created_users_count} usuario(s)<br>Actualizados: {updated_users_count} usuario(s)<br>"

def sheet_reader(sheet):
    wb = load_workbook(sheet)
    excel_data = list()

    try:
        emp_sheet = wb["empresas"]
        emp_res = reed_empresas(emp_sheet)
        excel_data.append(emp_res)
    except:
        pass
    try:
        art_sheet = wb["articulos"]
        art_res = reed_articulos(art_sheet)
        excel_data.append(art_res)
    except:
        pass
    try:
        usr_sheet = wb["usuarios"]
        usr_res = reed_usuarios(usr_sheet)
        excel_data.append(usr_res)
    except:
        pass
    
    return excel_data
    

def _get_actual_fields(model):
    fields = []
    for i in model._meta.get_fields():
        if (not isinstance(i, ManyToOneRel)):
            fields.append(i.name)
    return fields

def _write_header(sheet, fields):
    fill = PatternFill(
        fill_type="solid",
        start_color="9AC0CD",
        end_color="9AC0CD")
    for i in range(len(fields)):
        sheet.cell(1,i+1).value = fields[i].replace("_", " ").title()
        sheet.cell(1,i+1).fill = fill

def _readable_column_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    dims.get(cell.column_letter, 0), 
                    len(str(cell.value))+1
                )
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def _get_ret_names_list(retentions):
    for ret in retentions:
        yield ret[0]

def sheet_writer():
    wb = Workbook()

    # Create the sheets and write the headers
    art_sheet = wb.active
    art_sheet.title = "articulos"
    art_fields = _get_actual_fields(Articulo)
    print(art_fields)
    _write_header(art_sheet, art_fields)

    emp_sheet = wb.create_sheet(title="empresas")
    emp_fields = _get_actual_fields(Empresa)
    _write_header(emp_sheet, emp_fields)

    ing_sheet = wb.create_sheet(title="ingredientes")
    ing_fields = ('Nombre',)
    _write_header(ing_sheet, ing_fields)

    #write the data to the workbook
    row = 2
    for articulo in Articulo.objects.all():
        for field in range(len(art_fields)):
            cell = art_sheet.cell(row, field+1)
            if (art_fields[field] == "empresa"):
                cell.value = str(getattr(articulo, art_fields[field])) 
                continue
            cell.value = getattr(articulo, art_fields[field])
        row += 1

    row = 2
    for empresa in Empresa.objects.all():
        for field in range(len(emp_fields)):
            cell = emp_sheet.cell(row,field+1)
            if (emp_fields[field] == "retenciones"):
                cell.value = "+".join(
                    _get_ret_names_list(
                        empresa.retenciones.all().values_list('name')
                    )
                )
                continue
            cell.value = getattr(empresa, emp_fields[field])
        row += 1

    row = 2
    for ingrediente in Articulo.objects.all().values("ingrediente").distinct():
        cell = ing_sheet.cell(row,1)
        cell.value = ingrediente['ingrediente']
        row += 1

    # Data Validation
    empdv = DataValidation(
        type='list',
        formula1="{0}!${1}$2:${1}$3039".format(
            quote_sheetname(emp_sheet.title),
            get_column_letter(emp_fields.index("razon_social")+1))
        )
    art_sheet.add_data_validation(empdv)
    empdv.add("{0}2:{0}3039".format(
        get_column_letter(art_fields.index("empresa")+1))
    )

    ingdv = DataValidation(
        type='list',
        formula1="{0}!${1}$2:{1}$3039".format(
            quote_sheetname(ing_sheet.title),
            get_column_letter(ing_fields.index("Nombre")+1))
        )
    art_sheet.add_data_validation(ingdv)
    ingdv.add("{0}2:{0}3039".format(
        get_column_letter(art_fields.index("ingrediente")+1))
    )

    # Styling
    _readable_column_width(art_sheet)
    _readable_column_width(emp_sheet)
    _readable_column_width(ing_sheet)

    file = FileIO("dump.xlsx", 'w+')
    wb.save(file)
    file.seek(0)

    return file